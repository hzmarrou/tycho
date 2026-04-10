"""Tests for Excel export and gap report."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook


# ─── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_result():
    """A DataDictionaryResult with mixed-confidence elements."""
    from ontozense.extractors import (
        DataDictionaryResult,
        DataElement,
        FieldConfidence,
        Provenance,
    )

    result = DataDictionaryResult(
        domain_name="Non-Performing Loans",
        source_documents=["test-doc.md"],
        extraction_timestamp="2026-04-10T12:00:00",
    )

    # High-confidence element (everything filled)
    high = DataElement(
        element_name="Default",
        sub_domain="Loan",
        definition="A default occurs when an obligor is past due more than 90 days",
        term_definition="A default is considered to have occurred...",
        is_critical="Y",
        citation="Basel D403 §14",
        mandatory_optional="M",
        dq_completeness="Should be filled at facility level",
        dq_accuracy="Match Basel definition",
        dq_uniqueness="One per loan-borrower",
        dq_timeliness="Within 1 day",
        dq_consistency="Consistent across systems",
        dq_validity="Y/N",
    )
    high.confidence = [
        FieldConfidence("definition", 0.95, "verbatim"),
        FieldConfidence("sub_domain", 0.8, "non_empty"),
        FieldConfidence("is_critical", 0.8, "non_empty"),
        FieldConfidence("citation", 0.95, "verbatim"),
        FieldConfidence("dq_completeness", 0.8, "non_empty"),
    ]
    high.provenance = Provenance(source_document="test-doc.md", source_section="Section 3.1")
    result.elements.append(high)

    # Medium-confidence element (some empty)
    mid = DataElement(
        element_name="Forbearance",
        sub_domain="Forbearance",
        definition="A concession granted to a counterparty",
        is_critical="Y",
    )
    mid.confidence = [
        FieldConfidence("definition", 0.8, "non_empty"),
        FieldConfidence("sub_domain", 0.8, "non_empty"),
        FieldConfidence("is_critical", 0.8, "non_empty"),
        FieldConfidence("citation", 0.0, "empty"),
        FieldConfidence("dq_completeness", 0.0, "empty"),
        FieldConfidence("dq_accuracy", 0.0, "empty"),
    ]
    mid.provenance = Provenance(source_document="test-doc.md")
    result.elements.append(mid)

    # Low-confidence element (mostly empty)
    low = DataElement(
        element_name="Watch Status",
        sub_domain="",
        definition="",
    )
    low.confidence = [
        FieldConfidence("definition", 0.0, "empty"),
        FieldConfidence("sub_domain", 0.0, "empty"),
        FieldConfidence("is_critical", 0.0, "empty"),
    ]
    low.provenance = Provenance(source_document="test-doc.md")
    result.elements.append(low)

    return result


# ─── Excel exporter tests ────────────────────────────────────────────────────

class TestExcelExporter:
    def test_export_creates_file(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        exporter = DataDictionaryExcelExporter(sample_result)
        result_path = exporter.export(out)
        assert result_path.exists()
        assert result_path.stat().st_size > 0

    def test_main_sheet_has_correct_columns(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter
        from ontozense.exporters.excel import DD_COLUMNS, COMPUTED_COLUMNS

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out)

        wb = load_workbook(out)
        ws = wb["Data Dictionary"]
        # Headers are on row 3 (rows 1-2 are title and metadata)
        header_row = [c.value for c in ws[3]]
        # Standard DD columns first
        for label, _attr, _width in DD_COLUMNS:
            assert label in header_row
        # Then computed columns
        for label, _width in COMPUTED_COLUMNS:
            assert label in header_row

    def test_main_sheet_has_data_rows(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out)

        wb = load_workbook(out)
        ws = wb["Data Dictionary"]
        # Data starts on row 4 (header is row 3)
        # Read the "Data Element" column (index 2)
        names = [ws.cell(row=i, column=2).value for i in range(4, 4 + 3)]
        assert "Default" in names
        assert "Forbearance" in names
        assert "Watch Status" in names

    def test_confidence_column_populated(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter
        from ontozense.exporters.excel import DD_COLUMNS

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out)

        wb = load_workbook(out)
        ws = wb["Data Dictionary"]
        # Confidence column is right after the standard DD columns
        confidence_col = len(DD_COLUMNS) + 1
        # Default has high confidence
        default_conf = ws.cell(row=4, column=confidence_col).value
        assert "%" in str(default_conf)

    def test_needs_review_flag(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter
        from ontozense.exporters.excel import DD_COLUMNS, COMPUTED_COLUMNS

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out)

        wb = load_workbook(out)
        ws = wb["Data Dictionary"]
        # Needs Review column is the last computed column
        review_col = len(DD_COLUMNS) + len(COMPUTED_COLUMNS)
        # Read all needs-review values
        review_values = [ws.cell(row=i, column=review_col).value for i in range(4, 7)]
        # At least one should be Y (low/medium confidence elements)
        assert "Y" in review_values

    def test_gap_report_sheet_created(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out, include_gap_report=True)

        wb = load_workbook(out)
        assert "Gap Report" in wb.sheetnames

    def test_gap_report_disabled(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out, include_gap_report=False)

        wb = load_workbook(out)
        assert "Gap Report" not in wb.sheetnames

    def test_freeze_panes_set(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out)

        wb = load_workbook(out)
        ws = wb["Data Dictionary"]
        assert ws.freeze_panes is not None

    def test_auto_filter_applied(self, sample_result, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out)

        wb = load_workbook(out)
        ws = wb["Data Dictionary"]
        assert ws.auto_filter.ref is not None

    def test_conflicts_worksheet_created_when_conflicts_exist(self, tmp_path):
        from ontozense.exporters import DataDictionaryExcelExporter
        from ontozense.extractors import (
            DataDictionaryResult,
            DataElement,
            FieldConfidence,
            Provenance,
        )

        result = DataDictionaryResult(domain_name="Test Domain")
        el = DataElement(
            element_name="ItemA",
            definition="One definition from source A",
        )
        el.merge_conflicts = [
            "Source A says 'One definition'; Source B says 'Different definition'",
            "Source A flags as critical=Y; Source B flags as critical=N",
        ]
        el.confidence = [FieldConfidence("definition", 0.5, "non_empty")]
        el.provenance = Provenance(source_document="A.md")
        result.elements.append(el)

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(result).export(out, include_conflicts=True)

        wb = load_workbook(out)
        assert "Conflicts" in wb.sheetnames
        ws = wb["Conflicts"]
        # Header on row 3, conflict rows from row 4
        # Two conflicts → two rows
        assert ws.cell(row=4, column=1).value == "ItemA"
        assert ws.cell(row=5, column=1).value == "ItemA"

    def test_conflicts_worksheet_omitted_when_no_conflicts(self, sample_result, tmp_path):
        """If no conflicts exist, the Conflicts sheet should not be created."""
        from ontozense.exporters import DataDictionaryExcelExporter

        out = tmp_path / "test.xlsx"
        DataDictionaryExcelExporter(sample_result).export(out, include_conflicts=True)
        wb = load_workbook(out)
        # sample_result has no merge_conflicts on any element
        assert "Conflicts" not in wb.sheetnames


# ─── Gap report tests ────────────────────────────────────────────────────────

class TestGapReport:
    def test_compute_coverage(self, sample_result):
        from ontozense.exporters import compute_coverage

        coverage = compute_coverage(sample_result)
        # Should have one entry per data element field
        from ontozense.extractors.dd_extractor import DATA_ELEMENT_FIELDS
        assert len(coverage) == len(DATA_ELEMENT_FIELDS)

        # Definition is filled in 2 of 3 elements (Default and Forbearance)
        def_coverage = next(c for c in coverage if c[0] == "definition")
        assert def_coverage[1] == 2 / 3
        assert def_coverage[2] == 2  # filled
        assert def_coverage[3] == 1  # empty

    def test_generate_gap_report(self, sample_result):
        from ontozense.exporters import generate_gap_report

        report = generate_gap_report(sample_result, review_threshold=0.7)
        assert report.total_elements == 3
        assert report.high_confidence + report.medium_confidence + report.low_confidence == 3
        assert report.needs_review > 0  # at least one element needs review
        assert len(report.coverage_by_field) > 0
        assert len(report.suggested_actions) > 0

    def test_markdown_rendering(self, sample_result):
        from ontozense.exporters import generate_gap_report, render_to_markdown

        report = generate_gap_report(sample_result)
        md = render_to_markdown(report)
        assert "# Gap Report" in md
        assert "Non-Performing Loans" in md
        assert "Total elements" in md
        assert "Field Coverage" in md

    def test_save_markdown(self, sample_result, tmp_path):
        from ontozense.exporters import generate_gap_report, save_markdown

        report = generate_gap_report(sample_result)
        out = tmp_path / "gap.md"
        result_path = save_markdown(report, out)
        assert result_path.exists()
        content = result_path.read_text(encoding="utf-8")
        assert "# Gap Report" in content
