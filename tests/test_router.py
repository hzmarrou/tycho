"""Tests for the file router.

Covers all three layers per ``docs/PLAYBOOK.md`` §5:
  - Layer 1: deterministic file-extension rules + skip-list
  - Layer 2: content sniffing for ambiguous formats (.sql, .csv, .xlsx, .md)
  - Layer 3: NOT IMPLEMENTED YET (LLM classifier)

Tests use synthetic files with deliberately unambiguous content. Real
ambiguity is the failure mode the router exists to handle, so each test
constructs the minimum input that exercises one decision path.

Domain-neutral fixtures only — the router itself never sees domain content,
just structural patterns.
"""

from __future__ import annotations

from pathlib import Path

import pytest


# ─── Source enum + dataclass tests ───────────────────────────────────────────


class TestRoutingDecisionDataclass:
    def test_primary_source_returns_first(self):
        from ontozense.router import RoutingDecision, Source

        d = RoutingDecision(
            file_path=Path("x.md"),
            sources=[Source.A, Source.D],
            confidence=0.9,
            layer="content_sniff",
            reasoning="multi",
        )
        assert d.primary_source == Source.A

    def test_primary_source_empty_returns_skip(self):
        from ontozense.router import RoutingDecision, Source

        d = RoutingDecision(
            file_path=Path("x"),
            sources=[],
            confidence=0.0,
            layer="extension",
            reasoning="empty",
        )
        assert d.primary_source == Source.SKIP
        assert d.is_skip

    def test_is_multi_source_true_for_two_real_sources(self):
        from ontozense.router import RoutingDecision, Source

        d = RoutingDecision(
            file_path=Path("x.md"),
            sources=[Source.A, Source.D],
            confidence=0.9,
            layer="content_sniff",
            reasoning="multi",
        )
        assert d.is_multi_source

    def test_is_multi_source_false_for_one_real_source(self):
        from ontozense.router import RoutingDecision, Source

        d = RoutingDecision(
            file_path=Path("x.py"),
            sources=[Source.D],
            confidence=0.95,
            layer="extension",
            reasoning="single",
        )
        assert not d.is_multi_source

    def test_source_enum_values(self):
        from ontozense.router import Source

        assert Source.A.value == "A"
        assert Source.B.value == "B"
        assert Source.C.value == "C"
        assert Source.D.value == "D"
        assert Source.SKIP.value == "skip"


# ─── Layer 1 — extension rules ───────────────────────────────────────────────


class TestExtensionRules:
    @pytest.mark.parametrize(
        "filename,expected",
        [
            ("script.py", "D"),
            ("transform.dbt", "D"),
            ("model.r", "D"),
            ("Service.scala", "D"),
            ("Main.java", "D"),
        ],
    )
    def test_code_extensions_route_to_d(self, filename, expected, tmp_path):
        from ontozense.router import Router, Source

        path = tmp_path / filename
        decision = Router().route(path)
        assert decision.primary_source == Source(expected)
        assert decision.layer == "extension"

    @pytest.mark.parametrize(
        "filename",
        ["doc.md", "doc.markdown", "doc.txt", "doc.rst", "doc.pdf",
         "doc.docx", "doc.html"],
    )
    def test_prose_extensions_route_to_a(self, filename, tmp_path):
        from ontozense.router import Router, Source

        path = tmp_path / filename
        decision = Router().route(path)
        assert decision.primary_source == Source.A

    def test_ddl_extension_routes_to_c(self, tmp_path):
        from ontozense.router import Router, Source

        path = tmp_path / "schema.ddl"
        decision = Router().route(path)
        assert decision.primary_source == Source.C

    def test_unknown_extension_routes_to_skip(self, tmp_path):
        from ontozense.router import Router

        path = tmp_path / "weird.xyz"
        decision = Router().route(path)
        assert decision.is_skip
        assert "Unknown file extension" in decision.reasoning


# ─── Layer 1 — skip list ─────────────────────────────────────────────────────


class TestSkipList:
    @pytest.mark.parametrize(
        "filename",
        ["README.md", "readme.md", "README", "LICENSE", "LICENSE.txt",
         "CONTRIBUTING.md", "CHANGELOG.md", "CODE_OF_CONDUCT.md",
         ".gitignore", "Makefile", "Dockerfile"],
    )
    def test_skip_filenames(self, filename, tmp_path):
        from ontozense.router import Router

        path = tmp_path / filename
        decision = Router().route(path)
        assert decision.is_skip
        assert "skip list" in decision.reasoning

    @pytest.mark.parametrize(
        "filename",
        ["icon.png", "photo.jpg", "binary.exe", "lib.so", "archive.zip",
         "compiled.pyc", "debug.log"],
    )
    def test_skip_binary_extensions(self, filename, tmp_path):
        from ontozense.router import Router

        path = tmp_path / filename
        decision = Router().route(path)
        assert decision.is_skip
        assert "binary or build artifact" in decision.reasoning


# ─── Layer 2 — SQL content sniffing ──────────────────────────────────────────


class TestSqlContentSniff:
    def test_ddl_sql_routes_to_c(self, tmp_path):
        from ontozense.router import Router, Source

        sql_path = tmp_path / "schema.sql"
        sql_path.write_text(
            "-- Schema definition\n"
            "CREATE TABLE customer (id INT PRIMARY KEY, name VARCHAR(100));\n"
            "CREATE TABLE orders (id INT PRIMARY KEY, customer_id INT);\n",
            encoding="utf-8",
        )
        decision = Router().route(sql_path)
        assert decision.primary_source == Source.C
        assert decision.layer == "content_sniff"
        assert "DDL" in decision.reasoning

    def test_procedural_sql_routes_to_d(self, tmp_path):
        from ontozense.router import Router, Source

        sql_path = tmp_path / "logic.sql"
        sql_path.write_text(
            "CREATE FUNCTION compute_score(x INT)\n"
            "RETURNS INT AS $$\n"
            "DECLARE result INT;\n"
            "BEGIN\n"
            "  result := x * 2;\n"
            "  RETURN result;\n"
            "END;\n"
            "$$ LANGUAGE plpgsql;\n",
            encoding="utf-8",
        )
        decision = Router().route(sql_path)
        assert decision.primary_source == Source.D
        assert decision.layer == "content_sniff"
        assert "procedural" in decision.reasoning

    def test_empty_sql_falls_back_to_d(self, tmp_path):
        from ontozense.router import Router, Source

        sql_path = tmp_path / "noise.sql"
        sql_path.write_text("-- just a comment\n", encoding="utf-8")
        decision = Router().route(sql_path)
        assert decision.primary_source == Source.D


# ─── Layer 2 — CSV content sniffing ──────────────────────────────────────────


class TestCsvContentSniff:
    def test_governance_csv_routes_to_b(self, tmp_path):
        from ontozense.router import Router, Source

        csv_path = tmp_path / "dictionary.csv"
        csv_path.write_text(
            "Data Element,Definition,Critical Data Element,Mandatory/Optional,Completeness\n"
            "field_one,A first field,Y,M,Required\n",
            encoding="utf-8",
        )
        decision = Router().route(csv_path)
        assert decision.primary_source == Source.B
        assert decision.layer == "content_sniff"
        assert "governance" in decision.reasoning

    def test_schema_csv_routes_to_c(self, tmp_path):
        from ontozense.router import Router, Source

        csv_path = tmp_path / "schema_export.csv"
        csv_path.write_text(
            "Table Name,Column Name,Data Type,Nullable,Primary Key\n"
            "customer,id,INT,N,Y\n",
            encoding="utf-8",
        )
        decision = Router().route(csv_path)
        assert decision.primary_source == Source.C
        assert decision.layer == "content_sniff"
        assert "schema" in decision.reasoning

    def test_csv_unknown_headers_defaults_to_b(self, tmp_path):
        from ontozense.router import Router, Source

        csv_path = tmp_path / "data.csv"
        csv_path.write_text(
            "foo,bar,baz\n1,2,3\n4,5,6\n",
            encoding="utf-8",
        )
        decision = Router().route(csv_path)
        # Default for csv is B
        assert decision.primary_source == Source.B


# ─── Layer 2 — Excel content sniffing ────────────────────────────────────────


class TestExcelContentSniff:
    @pytest.fixture
    def governance_xlsx(self, tmp_path: Path) -> Path:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Data Element", "Definition", "Critical Data Element",
            "Mandatory / Optional", "Completeness", "Accuracy",
        ])
        ws.append(["field_one", "A first field", "Y", "M", "Required", "Verified"])
        path = tmp_path / "governance.xlsx"
        wb.save(path)
        return path

    @pytest.fixture
    def schema_xlsx(self, tmp_path: Path) -> Path:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Table Name", "Column Name", "Data Type",
            "Nullable", "Primary Key", "Foreign Key",
        ])
        ws.append(["customer", "id", "INT", "N", "Y", ""])
        path = tmp_path / "schema.xlsx"
        wb.save(path)
        return path

    @pytest.fixture
    def opaque_xlsx(self, tmp_path: Path) -> Path:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b", "c"])
        ws.append([1, 2, 3])
        path = tmp_path / "data.xlsx"
        wb.save(path)
        return path

    def test_governance_xlsx_routes_to_b(self, governance_xlsx):
        from ontozense.router import Router, Source

        decision = Router().route(governance_xlsx)
        assert decision.primary_source == Source.B
        assert decision.layer == "content_sniff"
        assert "governance" in decision.reasoning

    def test_schema_xlsx_routes_to_c(self, schema_xlsx):
        from ontozense.router import Router, Source

        decision = Router().route(schema_xlsx)
        assert decision.primary_source == Source.C
        assert decision.layer == "content_sniff"
        assert "schema" in decision.reasoning

    def test_opaque_xlsx_defaults_to_b(self, opaque_xlsx):
        from ontozense.router import Router, Source

        decision = Router().route(opaque_xlsx)
        # Default for ambiguous Excel is B
        assert decision.primary_source == Source.B


# ─── Layer 2 — Markdown content sniffing ─────────────────────────────────────


class TestMarkdownContentSniff:
    def test_pure_prose_markdown_routes_to_a(self, tmp_path):
        from ontozense.router import Router, Source

        md = tmp_path / "doc.md"
        md.write_text(
            "# A document\n\n"
            "## Section 1\n"
            "Some prose here describing things in detail.\n\n"
            "## Section 2\n"
            "More prose with no code at all.\n",
            encoding="utf-8",
        )
        decision = Router().route(md)
        assert decision.primary_source == Source.A
        assert not decision.is_multi_source

    def test_code_heavy_markdown_routes_to_a_and_d(self, tmp_path):
        from ontozense.router import Router, Source

        md = tmp_path / "guide.md"
        md.write_text(
            "# Developer guide\n\n"
            "## Setup\n\n"
            "```python\n"
            "import foo\n"
            "foo.bar()\n"
            "```\n\n"
            "## Usage\n\n"
            "```python\n"
            "result = foo.compute()\n"
            "```\n\n"
            "## Advanced\n\n"
            "```sql\n"
            "SELECT * FROM widgets;\n"
            "```\n",
            encoding="utf-8",
        )
        decision = Router().route(md)
        assert decision.is_multi_source
        assert Source.A in decision.sources
        assert Source.D in decision.sources
        assert decision.layer == "content_sniff"

    def test_markdown_with_one_code_block_still_just_a(self, tmp_path):
        from ontozense.router import Router, Source

        md = tmp_path / "mostly-prose.md"
        md.write_text(
            "# Document\n\nSome prose.\n\n```python\nfoo()\n```\n\nMore prose.\n",
            encoding="utf-8",
        )
        decision = Router().route(md)
        # 1 code block — not enough to also route to D
        assert decision.primary_source == Source.A
        assert not decision.is_multi_source


# ─── JSON content sniff ─────────────────────────────────────────────────────


class TestJsonContentSniff:
    """Governance JSON (element_name field) → Source B; JSON Schema → C."""

    def test_governance_json_array_routes_to_source_b(self, tmp_path):
        import json
        from ontozense.router import Router, Source

        f = tmp_path / "governance.json"
        f.write_text(
            json.dumps([
                {
                    "element_name": "Default",
                    "definition": "Inability to pay.",
                    "is_critical": True,
                },
                {"element_name": "Exposure"},
            ]),
            encoding="utf-8",
        )
        decision = Router().route(f)
        assert decision.primary_source == Source.B
        assert decision.layer == "content_sniff"
        assert decision.confidence >= 0.9

    def test_governance_json_single_object_routes_to_source_b(self, tmp_path):
        import json
        from ontozense.router import Router, Source

        f = tmp_path / "gov.json"
        f.write_text(
            json.dumps({
                "element_name": "Default",
                "definition": "A status.",
            }),
            encoding="utf-8",
        )
        decision = Router().route(f)
        assert decision.primary_source == Source.B

    def test_json_schema_routes_to_source_c(self, tmp_path):
        import json
        from ontozense.router import Router, Source

        f = tmp_path / "schema.json"
        f.write_text(
            json.dumps({
                "$schema": "http://json-schema.org/draft-07/schema#",
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                },
            }),
            encoding="utf-8",
        )
        decision = Router().route(f)
        assert decision.primary_source == Source.C
        assert decision.layer == "content_sniff"

    def test_openapi_spec_routes_to_source_c(self, tmp_path):
        import json
        from ontozense.router import Router, Source

        f = tmp_path / "api.json"
        f.write_text(
            json.dumps({
                "openapi": "3.0.0",
                "info": {"title": "Test API"},
                "paths": {},
            }),
            encoding="utf-8",
        )
        decision = Router().route(f)
        assert decision.primary_source == Source.C

    def test_avro_schema_routes_to_source_c(self, tmp_path):
        import json
        from ontozense.router import Router, Source

        f = tmp_path / "event.json"
        f.write_text(
            json.dumps({
                "type": "record",
                "name": "Event",
                "fields": [
                    {"name": "id", "type": "string"},
                ],
            }),
            encoding="utf-8",
        )
        decision = Router().route(f)
        assert decision.primary_source == Source.C

    def test_unknown_json_falls_back_to_c(self, tmp_path):
        """A generic JSON without governance/schema markers: Source C
        with lower confidence (invite human review)."""
        import json
        from ontozense.router import Router, Source

        f = tmp_path / "data.json"
        f.write_text(
            json.dumps({"some": "random", "data": [1, 2, 3]}),
            encoding="utf-8",
        )
        decision = Router().route(f)
        assert decision.primary_source == Source.C
        # Below the --auto threshold (0.9), so --auto would skip it
        assert decision.confidence < 0.9

    def test_malformed_json_falls_back_to_c(self, tmp_path):
        from ontozense.router import Router, Source

        f = tmp_path / "broken.json"
        f.write_text("this is { not valid json", encoding="utf-8")
        decision = Router().route(f)
        assert decision.primary_source == Source.C

    def test_real_governance_example_fixture_routes_to_b(self):
        """The shipped docs/governance_example.json should route to B."""
        from pathlib import Path
        from ontozense.router import Router, Source

        fixture = Path("docs/governance_example.json")
        if not fixture.exists():
            import pytest
            pytest.skip("governance_example.json not found")

        decision = Router().route(fixture)
        assert decision.primary_source == Source.B, (
            f"Shipped governance fixture should route to B, got "
            f"{decision.primary_source.value}. This is the tutorial's "
            f"canonical example."
        )


# ─── Directory routing ──────────────────────────────────────────────────────


class TestDirectoryRouting:
    def test_route_directory_returns_one_decision_per_file(self, tmp_path):
        from ontozense.router import Router

        (tmp_path / "doc.md").write_text("# Doc\n", encoding="utf-8")
        (tmp_path / "code.py").write_text("def foo(): pass\n", encoding="utf-8")
        (tmp_path / "readme.md").write_text("# Readme\n", encoding="utf-8")

        decisions = Router().route_directory(tmp_path)
        assert len(decisions) == 3

    def test_route_directory_recursive(self, tmp_path):
        from ontozense.router import Router

        (tmp_path / "top.md").write_text("# Top\n", encoding="utf-8")
        sub = tmp_path / "sub"
        sub.mkdir()
        (sub / "deep.py").write_text("x = 1\n", encoding="utf-8")

        decisions = Router().route_directory(tmp_path, recursive=True)
        # Both files should be discovered
        names = [d.file_path.name for d in decisions]
        assert "top.md" in names
        assert "deep.py" in names

    def test_route_directory_non_recursive(self, tmp_path):
        from ontozense.router import Router

        (tmp_path / "top.md").write_text("# Top\n", encoding="utf-8")
        sub = tmp_path / "sub"
        sub.mkdir()
        (sub / "deep.py").write_text("x = 1\n", encoding="utf-8")

        decisions = Router().route_directory(tmp_path, recursive=False)
        names = [d.file_path.name for d in decisions]
        assert "top.md" in names
        assert "deep.py" not in names

    def test_route_directory_raises_on_file(self, tmp_path):
        from ontozense.router import Router

        f = tmp_path / "x.md"
        f.write_text("foo", encoding="utf-8")
        with pytest.raises(NotADirectoryError):
            Router().route_directory(f)


# ─── Convenience function ────────────────────────────────────────────────────


class TestRouteFileHelper:
    def test_route_file_returns_decision(self, tmp_path):
        from ontozense.router import route_file, Source

        path = tmp_path / "script.py"
        decision = route_file(path)
        assert decision.primary_source == Source.D
