"""Excel exporter for data dictionary results.

Outputs an Excel file with the column structure used by typical enterprise
data dictionaries (Domain, Data Element, Definition, Critical flag,
Citation, Mandatory/Optional, and the standard six data quality dimensions).
Format is intentionally familiar to expert reviewers — they open the file
in Excel and review in their existing workflow.

Adds three columns beyond the standard data dictionary format:
  - Confidence (red/yellow/green by aggregate score)
  - Source Documents (provenance)
  - Needs Review (Y/N — flagged for human attention)
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..extractors.dd_extractor import DataDictionaryResult, DataElement


# ─── Column definitions ──────────────────────────────────────────────────────

# (header_label, DataElement attribute name, column width)
# Note: the header is "Sub-domain" not "Domain". The document-level domain
# (the high-level area covered by the source document) appears in the title
# row above the table. The Sub-domain column is per-element categorization
# within that domain.
DD_COLUMNS: list[tuple[str, str, int]] = [
    ("Sub-domain", "sub_domain", 20),
    ("Data Element", "element_name", 30),
    ("Data Element Definition", "definition", 60),
    ("Critical Data Element", "is_critical", 12),
    ("Citation", "citation", 30),
    ("Term Definition (formal)", "term_definition", 60),
    ("Mandatory / Optional", "mandatory_optional", 12),
    ("Completeness", "dq_completeness", 40),
    ("Accuracy", "dq_accuracy", 30),
    ("Uniqueness", "dq_uniqueness", 30),
    ("Timeliness", "dq_timeliness", 25),
    ("Consistency", "dq_consistency", 30),
    ("Validity", "dq_validity", 30),
]

# Computed columns appended after the standard columns
COMPUTED_COLUMNS: list[tuple[str, int]] = [
    ("Confidence", 12),
    ("Source Documents", 40),
    ("Needs Review", 14),
]


# ─── Styling ─────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
COMPUTED_HEADER_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

CONFIDENCE_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
CONFIDENCE_MID = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # yellow
CONFIDENCE_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red

NEEDS_REVIEW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ─── Exporter ────────────────────────────────────────────────────────────────

class DataDictionaryExcelExporter:
    """Exports a DataDictionaryResult to an Excel workbook."""

    def __init__(self, result: DataDictionaryResult):
        self.result = result

    def export(
        self,
        output_path: str | Path,
        include_gap_report: bool = True,
        include_conflicts: bool = True,
        review_threshold: float = 0.7,
    ) -> Path:
        """Write the data dictionary to an Excel file.

        Args:
            output_path: Where to save the .xlsx file
            include_gap_report: Add a "Gap Report" worksheet
            include_conflicts: Add a "Conflicts" worksheet (if any conflicts exist)
            review_threshold: Confidence threshold below which an element is flagged for review

        Returns:
            Path to the written file
        """
        output_path = Path(output_path)
        wb = Workbook()

        # Main Data Dictionary sheet
        ws = wb.active
        ws.title = "Data Dictionary"
        self._write_dd_sheet(ws, review_threshold)

        if include_gap_report:
            gap_sheet = wb.create_sheet("Gap Report")
            self._write_gap_sheet(gap_sheet, review_threshold)

        if include_conflicts:
            conflicts = [(el, c) for el in self.result.elements for c in el.merge_conflicts]
            if conflicts:
                conflict_sheet = wb.create_sheet("Conflicts")
                self._write_conflicts_sheet(conflict_sheet, conflicts)

        wb.save(output_path)
        return output_path

    # ─── Main DD sheet ────────────────────────────────────────────────────

    def _write_dd_sheet(self, ws: Worksheet, review_threshold: float) -> None:
        n_cols = len(DD_COLUMNS) + len(COMPUTED_COLUMNS)

        # Title row
        ws.cell(row=1, column=1, value="Data Dictionary")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

        # Domain label (clearly visible — separate from the title)
        domain_label = self.result.domain_name or "(unspecified)"
        ws.cell(row=2, column=1, value="Document Domain:").font = Font(bold=True, color="1F4E79")
        ws.cell(row=2, column=2, value=domain_label).font = Font(bold=True)

        ws.cell(row=2, column=4, value="Generated:").font = Font(bold=True, color="1F4E79")
        ws.cell(row=2, column=5, value=self.result.extraction_timestamp).font = Font(italic=True, color="666666")

        ws.cell(row=2, column=8, value="Sources:").font = Font(bold=True, color="1F4E79")
        sources_cell = ws.cell(
            row=2, column=9,
            value=", ".join(Path(s).name for s in self.result.source_documents),
        )
        sources_cell.font = Font(italic=True, color="666666")

        # Header row (row 3)
        header_row = 3
        col = 1
        for label, _attr, width in DD_COLUMNS:
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
            col += 1

        for label, width in COMPUTED_COLUMNS:
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.font = HEADER_FONT
            cell.fill = COMPUTED_HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
            col += 1

        ws.row_dimensions[header_row].height = 30

        # Freeze header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Data rows
        for i, element in enumerate(self.result.elements, start=header_row + 1):
            self._write_element_row(ws, i, element, review_threshold)

        # Auto-filter on the header row
        last_col = get_column_letter(len(DD_COLUMNS) + len(COMPUTED_COLUMNS))
        ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(self.result.elements)}"

    def _write_element_row(
        self,
        ws: Worksheet,
        row: int,
        element: DataElement,
        review_threshold: float,
    ) -> None:
        col = 1
        for _label, attr, _width in DD_COLUMNS:
            value = getattr(element, attr, "")
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            col += 1

        # Confidence column
        conf = element.overall_confidence()
        conf_cell = ws.cell(row=row, column=col, value=f"{conf:.0%}")
        conf_cell.alignment = Alignment(horizontal="center", vertical="top")
        if conf >= 0.8:
            conf_cell.fill = CONFIDENCE_HIGH
        elif conf >= 0.5:
            conf_cell.fill = CONFIDENCE_MID
        else:
            conf_cell.fill = CONFIDENCE_LOW
        col += 1

        # Source Documents column
        sources = ""
        if element.provenance:
            source_name = Path(element.provenance.source_document).name
            sources = source_name
        ws.cell(row=row, column=col, value=sources).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1

        # Needs Review column
        needs_review = element.needs_review(threshold=review_threshold)
        review_cell = ws.cell(row=row, column=col, value="Y" if needs_review else "N")
        review_cell.alignment = Alignment(horizontal="center", vertical="top")
        if needs_review:
            review_cell.fill = NEEDS_REVIEW_FILL

    # ─── Gap Report sheet ─────────────────────────────────────────────────

    def _write_gap_sheet(self, ws: Worksheet, review_threshold: float) -> None:
        ws.cell(row=1, column=1, value="Gap Report").font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Summary stats
        total = len(self.result.elements)
        high = sum(1 for el in self.result.elements if el.overall_confidence() >= 0.8)
        mid = sum(1 for el in self.result.elements if 0.5 <= el.overall_confidence() < 0.8)
        low = sum(1 for el in self.result.elements if el.overall_confidence() < 0.5)
        review = sum(1 for el in self.result.elements if el.needs_review(review_threshold))

        ws.cell(row=3, column=1, value="Total elements:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=total)
        ws.cell(row=4, column=1, value="High confidence (≥80%):").font = Font(bold=True)
        ws.cell(row=4, column=2, value=high)
        ws.cell(row=5, column=1, value="Medium confidence (50-79%):").font = Font(bold=True)
        ws.cell(row=5, column=2, value=mid)
        ws.cell(row=6, column=1, value="Low confidence (<50%):").font = Font(bold=True)
        ws.cell(row=6, column=2, value=low)
        ws.cell(row=7, column=1, value="Needs human review:").font = Font(bold=True)
        ws.cell(row=7, column=2, value=review)

        # Field coverage stats
        ws.cell(row=9, column=1, value="Field Coverage").font = Font(bold=True, size=12)

        coverage_headers = ["Field", "% Filled", "Filled", "Empty"]
        for i, h in enumerate(coverage_headers, start=1):
            cell = ws.cell(row=10, column=i, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        from .gap_report import compute_coverage  # local import to avoid cycle issues
        coverage = compute_coverage(self.result)
        for i, (field_name, pct, filled, empty) in enumerate(coverage, start=11):
            ws.cell(row=i, column=1, value=field_name)
            ws.cell(row=i, column=2, value=f"{pct:.0%}")
            ws.cell(row=i, column=3, value=filled)
            ws.cell(row=i, column=4, value=empty)
            # Color code
            row_fill = (
                CONFIDENCE_HIGH if pct >= 0.8
                else CONFIDENCE_MID if pct >= 0.5
                else CONFIDENCE_LOW
            )
            ws.cell(row=i, column=2).fill = row_fill

        for col_idx, width in enumerate([35, 12, 10, 10], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ─── Conflicts sheet ──────────────────────────────────────────────────

    def _write_conflicts_sheet(self, ws: Worksheet, conflicts: list[tuple]) -> None:
        ws.cell(row=1, column=1, value="Merge Conflicts").font = Font(bold=True, size=14)
        ws.merge_cells("A1:C1")

        headers = ["Element", "Conflict Description", "Action Required"]
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for i, (element, conflict_desc) in enumerate(conflicts, start=4):
            ws.cell(row=i, column=1, value=element.element_name)
            ws.cell(row=i, column=2, value=conflict_desc)
            ws.cell(row=i, column=3, value="Review and resolve")

        for col_idx, width in enumerate([30, 60, 25], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
