"""Excel exporter for Source A (domain document) extraction results.

Outputs a per-source Excel workbook with two sheets:

  - **Concepts** — concept name, definition, citation, source, confidence,
    needs review
  - **Relationships** — subject, predicate, object, source, confidence

This is the **per-source view** — one Excel per Source A run, showing
exactly what that single source contributed. The rich 13-field data
dictionary is the OUTPUT of the fusion layer (Step 6), assembled from all
four sources. The two views serve different review purposes:

  - Per-source: "what did this document tell us?"
  - Fusion-layer rich: "what is the complete data dictionary for this
    domain, with provenance per field?"

Both are written to disk so the human reviewer can inspect either lens.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..extractors.domain_doc_extractor import DomainDocumentExtractionResult


# ─── Column definitions ──────────────────────────────────────────────────────


CONCEPT_COLUMNS: list[tuple[str, int]] = [
    ("Concept", 30),
    ("Definition", 60),
    ("Citation", 25),
    ("Confidence", 12),
    ("Source Document", 30),
    ("Source Section", 25),
    ("Needs Review", 14),
]

RELATIONSHIP_COLUMNS: list[tuple[str, int]] = [
    ("Subject", 30),
    ("Predicate", 25),
    ("Object", 30),
    ("Confidence", 12),
    ("Source Document", 30),
]


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

CONFIDENCE_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CONFIDENCE_MID = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CONFIDENCE_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

NEEDS_REVIEW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ─── Exporter ────────────────────────────────────────────────────────────────


class DomainDocumentExcelExporter:
    """Writes a Source A extraction result to a 2-sheet Excel workbook."""

    def __init__(self, result: DomainDocumentExtractionResult):
        self.result = result

    def export(
        self,
        output_path: str | Path,
        review_threshold: float = 0.7,
    ) -> Path:
        """Write the extraction result to an Excel file at ``output_path``."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        concepts_sheet = wb.active
        concepts_sheet.title = "Concepts"
        self._write_concepts_sheet(concepts_sheet, review_threshold)

        rels_sheet = wb.create_sheet("Relationships")
        self._write_relationships_sheet(rels_sheet)

        wb.save(output_path)
        return output_path

    # ─── Concepts sheet ───────────────────────────────────────────────────

    def _write_concepts_sheet(self, ws: Worksheet, review_threshold: float) -> None:
        n_cols = len(CONCEPT_COLUMNS)

        # Title and metadata rows
        ws.cell(row=1, column=1, value="Source A — Domain Document Extraction (Concepts)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

        ws.cell(row=2, column=1, value="Document Domain:").font = Font(bold=True, color="1F4E79")
        ws.cell(row=2, column=2, value=self.result.domain_name or "(unspecified)").font = Font(bold=True)

        ws.cell(row=2, column=4, value="Generated:").font = Font(bold=True, color="1F4E79")
        ws.cell(row=2, column=5, value=self.result.extraction_timestamp).font = Font(italic=True, color="666666")

        # Header row
        header_row = 3
        for col_idx, (label, width) in enumerate(CONCEPT_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[header_row].height = 30
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Data rows
        for i, concept in enumerate(self.result.concepts, start=header_row + 1):
            self._write_concept_row(ws, i, concept, review_threshold)

        # Auto-filter
        if self.result.concepts:
            last_col = get_column_letter(n_cols)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(self.result.concepts)}"

    def _write_concept_row(
        self,
        ws: Worksheet,
        row: int,
        concept,
        review_threshold: float,
    ) -> None:
        col = 1
        # Concept
        ws.cell(row=row, column=col, value=concept.name).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        # Definition
        ws.cell(row=row, column=col, value=concept.definition).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        # Citation
        ws.cell(row=row, column=col, value=concept.citation).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        # Confidence
        conf = concept.overall_confidence()
        conf_cell = ws.cell(row=row, column=col, value=f"{conf:.0%}")
        conf_cell.alignment = Alignment(horizontal="center", vertical="top")
        if conf >= 0.8:
            conf_cell.fill = CONFIDENCE_HIGH
        elif conf >= 0.5:
            conf_cell.fill = CONFIDENCE_MID
        else:
            conf_cell.fill = CONFIDENCE_LOW
        col += 1
        # Source document
        source = ""
        if concept.provenance:
            source = Path(concept.provenance.source_document).name
        ws.cell(row=row, column=col, value=source).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        # Source section
        section = ""
        if concept.provenance:
            section = concept.provenance.source_section
        ws.cell(row=row, column=col, value=section).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        # Needs review
        needs_review = concept.needs_review(threshold=review_threshold)
        review_cell = ws.cell(row=row, column=col, value="Y" if needs_review else "N")
        review_cell.alignment = Alignment(horizontal="center", vertical="top")
        if needs_review:
            review_cell.fill = NEEDS_REVIEW_FILL

    # ─── Relationships sheet ──────────────────────────────────────────────

    def _write_relationships_sheet(self, ws: Worksheet) -> None:
        n_cols = len(RELATIONSHIP_COLUMNS)

        ws.cell(row=1, column=1, value="Source A — Domain Document Extraction (Relationships)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

        header_row = 3
        for col_idx, (label, width) in enumerate(RELATIONSHIP_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[header_row].height = 30
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        for i, rel in enumerate(self.result.relationships, start=header_row + 1):
            self._write_relationship_row(ws, i, rel)

        if self.result.relationships:
            last_col = get_column_letter(n_cols)
            ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(self.result.relationships)}"

    def _write_relationship_row(self, ws: Worksheet, row: int, rel) -> None:
        col = 1
        ws.cell(row=row, column=col, value=rel.subject).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        ws.cell(row=row, column=col, value=rel.predicate).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        ws.cell(row=row, column=col, value=rel.object).alignment = Alignment(vertical="top", wrap_text=True)
        col += 1
        conf = rel.overall_confidence()
        conf_cell = ws.cell(row=row, column=col, value=f"{conf:.0%}")
        conf_cell.alignment = Alignment(horizontal="center", vertical="top")
        if conf >= 0.8:
            conf_cell.fill = CONFIDENCE_HIGH
        elif conf >= 0.5:
            conf_cell.fill = CONFIDENCE_MID
        else:
            conf_cell.fill = CONFIDENCE_LOW
        col += 1
        source = ""
        if rel.provenance:
            source = Path(rel.provenance.source_document).name
        ws.cell(row=row, column=col, value=source).alignment = Alignment(vertical="top", wrap_text=True)
